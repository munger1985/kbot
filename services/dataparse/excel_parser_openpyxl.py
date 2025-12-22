import json
import uuid
import zipfile
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import shutil
from loguru import logger
from .file_params import FileParams
from .common import check_text_file, save_embeddings
from dao.repositories.kbot_md_kb_files_repo import KbotMdKbFilesRepository
from dao.entities.kbot_biz_txt_embedding import KbotBizTxtEmbedding
from core.dictionary import FileStatus, ChunkType, SplitStrategy
from utils.call_models import CallModel
from .summary_parser import SummaryParser

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
except ImportError:
    print("请安装openpyxl: pip install openpyxl")
    exit(1)

try:
    from PIL import Image as PILImage
    import io
except ImportError:
    print("请安装Pillow: pip install Pillow")
    exit(1)


class ExcelParser:
    def __init__(self, file_params: FileParams):
        self.file_params = file_params
        self.parsed_metadata = ''

        """
        初始化Excel图片和数据提取器

        Args:
            excel_file_path: Excel文件路径
            output_dir: 输出目录
        """
        self.excel_file_path = Path(file_params.file_path)
        self.output_dir = self.excel_file_path.parent / "output"/file_params.file_id
        self.images_dir = self.output_dir / "images"
        self.data_dir = self.output_dir / "data"
        self.chunk_size = 0
        self.chunk_overlap = 0
        self.images_info: list[dict] = []
        self.tables_info: Dict[str, Any] = {}
        # 创建输出目录
        self.output_dir.mkdir(parents=True,exist_ok=True)
        self.images_dir.mkdir(parents=True,exist_ok=True)
        self.data_dir.mkdir(parents=True,exist_ok=True)

        # 存储提取的数据
        self.extracted_data = {}

    def generate_random_image_id(self) -> str:
        """生成随机图片ID"""
        return str(uuid.uuid4())

    async def _update_file_status(self, status: FileStatus, message: str) -> None:
        """Helper method to update file status"""
        await KbotMdKbFilesRepository().update_file_status(
            self.file_params.file_id,
            status,
            message
        )
    # async def _save_embeddings(self, embeddings: list[KbotBizTxtEmbedding]) -> bool:
    #     """Save embeddings to database with error handling"""
    #     if not embeddings:
    #         return False

    #     try:
    #         repo = KbotBizTxtEmbeddingRepository(kb_id=self.file_params.kb_id)
    #         await repo.initialize()
    #         result = await repo.create(kb_id=self.file_params.kb_id, embeddings=embeddings)
    #         if not result:
    #             msg = "Failed to save embeddings (repository returned False)"
    #             logger.error(msg)
    #             await self._update_file_status(FileStatus.PARSE_FAILED, msg)
    #             return False

    #         logger.info(f"Successfully saved {len(embeddings)} embeddings")
    #         return True

    #     except Exception as e:
    #         msg = f"Exception while saving embeddings: {str(e)}"
    #         logger.error(msg)
    #         await self._update_file_status(FileStatus.PARSE_FAILED, msg)
    #         return False
    def get_sheet_naming_convention(self, sheet_name: str) -> str:
        """
        根据工作表名称应用命名约定

        Args:
            sheet_name: 原始工作表名称

        Returns:
            格式化后的工作表名称
        """
        # 移除特殊字符，替换空格为下划线
        clean_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_'))
        clean_name = clean_name.replace(' ', '_').replace('-', '_')

        # 添加时间戳前缀确保唯一性
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{timestamp}_{clean_name}"

    def extract_images_from_sheet(self, sheet_name: str, sheet_data: Dict) -> List[Dict]:
        """
        从工作表中提取图片

        Args:
            sheet_name: 工作表名称
            sheet_data: 工作表数据

        Returns:
            图片信息列表
        """
        images = []

        try:
            # 加载工作簿
            workbook = load_workbook(self.excel_file_path, data_only=True)
            worksheet = workbook[sheet_name]

            # 检查是否有图片
            if hasattr(worksheet, '_images'):
                for idx, img in enumerate(worksheet._images): # type: ignore
                    try:
                        # 生成随机图片ID
                        image_id = self.generate_random_image_id()

                        # 保存图片文件
                        image_filename = f"{image_id}.png"
                        image_path = (self.images_dir / image_filename).absolute()

                        # 将图片数据保存到文件
                        with open(image_path, 'wb') as f:
                            f.write(img._data())

                        # 记录图片信息
                        image_info = {
                            "image_id": image_id,
                            "filename": image_filename,
                            "sheet_name": sheet_name,
                            "position": {
                                "row": img.anchor._from.row if hasattr(img.anchor, '_from') else None,
                                "col": img.anchor._from.col if hasattr(img.anchor, '_from') else None
                            },
                            "size": {
                                "width": img.width if hasattr(img, 'width') else None,
                                "height": img.height if hasattr(img, 'height') else None
                            },
                            "extracted_at": datetime.now().isoformat()
                        }

                        images.append(image_info)

                    except Exception as e:
                        logger.error(f"提取图片时出错: {e}")
                        continue

            # 尝试从Excel的压缩文件中提取图片（适用于某些Excel版本）
            if not images:
                images.extend(self._extract_images_from_zip(sheet_name))

        except Exception as e:
            logger.error(f"从工作表 {sheet_name} 提取图片时出错: {e}")

        return images

    def extract_images_metadata(self, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        提取图片并返回元数据列表（仅包含sheet与image_id等必要信息）

        Args:
            sheet_name: 指定工作表名称；为None时处理所有工作表

        Returns:
            图片元数据列表
        """
        metadata: List[Dict[str, Any]] = []
        try:
            workbook = load_workbook(self.excel_file_path, data_only=True)
            target_sheets = [sheet_name] if sheet_name else workbook.sheetnames
            for current_sheet in target_sheets:
                try:
                    worksheet = workbook[current_sheet]
                except Exception:
                    continue
                # 仅使用 openpyxl 的 _images 获取该工作表内的图片，避免ZIP方式的跨表误关联
                if hasattr(worksheet, '_images') and worksheet._images: # type: ignore
                    for img in worksheet._images: # type: ignore
                        try:
                            image_id = self.generate_random_image_id()
                            image_filename = f"{image_id}.png"
                            image_path = (self.images_dir / image_filename).absolute()

                            # 保存图片为文件
                            with open(image_path, 'wb') as f:
                                f.write(img._data())
                            metadata.append({
                                "sheet_name": current_sheet,
                                "image_id": image_id,
                                "file_path": str(image_path)
                            })
                        except Exception:
                            continue
        except Exception as e:
            logger.error(f"提取图片元数据时出错: {e}")
        return metadata

    def save_image_metadata(self, metadata: List[Dict[str, Any]], filename: str = "image_info.json") -> str:
        """
        保存图片元数据到JSON文件

        Args:
            metadata: 图片元数据列表
            filename: 输出文件名（默认 image_metadata.json）

        Returns:
            保存的文件路径
        """
        if not metadata:
            logger.info("没有图片元数据可保存")
            return ""
        json_path = self.data_dir / filename
        try:
            json_str = json.dumps({"images": metadata}, ensure_ascii=False, indent=2)
            with open("output.json", "w", encoding="utf-8") as f:
                f.write(json_str)
            logger.info(f"图片元数据已保存到: {json_path.absolute()}")
            return json_str
        except Exception as e:
            logger.error(f"保存图片元数据时出错: {e}")
            return ""

    def __await__(self):
        # 返回一个生成器或迭代器
        yield  # 模拟异步操作
        return "自定义协程结果"
    def extract_images_and_save_metadata(self, sheet_name: Optional[str] = None,
                                         filename: str = "image_info.json") -> str:
        """
        便捷方法：提取图片并直接保存元数据到文件（默认 image_info.json）

        Args:
            sheet_name: 指定工作表名称；为None时处理所有工作表
            filename: 输出文件名

        Returns:
            保存的文件路径
        """
        metadata = self.extract_images_metadata(sheet_name)
        self.images_info= metadata
        return self.save_image_metadata(metadata, filename)

    async def _process_embeddings(self) -> bool:
        """Process text and table embeddings for by fixed size"""
        
        if not self.file_params.txt_embed_model:
            msg = f"Embedding model not found for id: {self.file_params.txt_embed_model}"
            logger.error(msg)
            await self._update_file_status(FileStatus.PARSE_FAILED, msg)
            return False

        # Prepare all content chunks for embedding
        chunks = []
        chunk_metas = []

        # Add table content
        for sheet in self.tables_info:
                sheet_data = self.tables_info[sheet]
                rows =sheet_data['data']
                for row in rows:
                    chunks.append(json.dumps(row,ensure_ascii=False))
                    chunk_metas.append({
                        "chunk_type": ChunkType.TABLE,
                        "sheet_name": sheet
                    })



        if not chunks:
            logger.warning("No valid content chunks found for embedding")
            return True  # Consider empty content as success

        # Get all embeddings in one call
        embeddings_list = await CallModel().call_embedding_model(self.file_params.txt_embed_model, chunks)
        if not embeddings_list or len(embeddings_list) != len(chunks):
            msg = f"Embedding model {self.file_params.txt_embed_model} returned invalid results (expected {len(chunks)}, got {len(embeddings_list) if embeddings_list else 0})"
            logger.error(msg)
            await self._update_file_status(FileStatus.PARSE_FAILED, msg)
            return False

        # Create embedding entities
        embed_entities = []
        for idx, (chunk, meta) in enumerate(zip(chunks, chunk_metas)):
            embed_entity = KbotBizTxtEmbedding(
                kb_id=self.file_params.kb_id,
                embed_id=str(uuid.uuid4()),
                chunk_doc=chunk,
                chunk_metadata=meta,
                biz_metadata=self.file_params.biz_metadata,
                file_id=self.file_params.file_id,
                embedding=embeddings_list[idx].embedding,
                security_level=self.file_params.security_level,
                status=1
            )
            embed_entities.append(embed_entity)
        image_embed_entities= await self._process_images_embeddings()
        embed_entities.extend(image_embed_entities)

        if self.file_params.enable_summary:
            logger.debug("启用摘要处理")
            summary_result = await SummaryParser.process_summary(file_params=self.file_params, embed_entities=embed_entities)
            # return summary_result

        # Save all embeddings in one batch
        return await save_embeddings(file_params=self.file_params, embeddings=embed_entities)

    async def _process_images_embeddings(self) -> list:
        ## 1 means yes
        if self.file_params.img2txt == 1:
            if self.file_params.img2txt_model is None:
                msg = f"Image to text model not found for id: {self.file_params.img2txt_model}"
                logger.error(msg)
                await self._update_file_status(FileStatus.PARSE_FAILED, msg)
                return []
            
            self.parsed_metadata = self.extract_images_and_save_metadata(filename="image_info.json")

            # if self.file_params.parser.get("extract_images", False):
            vlm_prompt_unique_name = "SYSTEM/image2text"
            
            chunks = []
            chunk_metas = []

            for eachImage in self.images_info:
                description_file = Path(eachImage['file_path'] + ".description")
                if not description_file.exists():

                    image_description = await CallModel().call_vlm_model_for_parsing_picture(self.file_params.img2txt_model,
                                                                                             vlm_prompt_unique_name,
                                                                                             eachImage['file_path'])
                    if image_description:
                        description_file.write_text(
                            image_description,
                            encoding='utf-8'
                        )
                        chunk_metas.append({
                            "chunk_type": ChunkType.IMAGE,
                            "sheet_name": eachImage['sheet_name'],
                            "image_id": eachImage['image_id'],
                        })
                        chunks.append(image_description)

            if not self.file_params.txt_embed_model:
                msg = f"text_embedding_model not found for id: {self.file_params.txt_embed_model}"
                logger.error(msg)
                await self._update_file_status(FileStatus.PARSE_FAILED, msg)
                return []
            embeddings_list = []
            if chunks:
                embeddings_list = await CallModel().call_embedding_model(self.file_params.txt_embed_model, chunks)
            if embeddings_list and len(embeddings_list) != len(chunks):
                msg = f"text_embedding_model  {self.file_params.txt_embed_model} returned invalid results (expected {len(chunks)}, got {len(embeddings_list) if embeddings_list else 0})"
                logger.error(msg)
                logger.error("failed file: {}", self.file_params.file_path)
                await self._update_file_status(FileStatus.PARSE_FAILED, msg)
                return []

                # Create embedding entities
            embed_entities = []
            for idx, (chunk, meta) in enumerate(zip(chunks, chunk_metas)):
                embed_entity = KbotBizTxtEmbedding(
                    kb_id=self.file_params.kb_id,
                    embed_id=meta['image_id'],
                    chunk_doc=chunk,
                    chunk_metadata=meta,
                    biz_metadata=self.file_params.biz_metadata,
                    file_id=self.file_params.file_id,
                    embedding=embeddings_list[idx].embedding,  # type: ignore
                    security_level=self.file_params.security_level,
                    status=1
                )
                embed_entities.append(embed_entity)

            # Save all embeddings in one batch
            return embed_entities
        else:
            return []

    def _extract_images_from_zip(self, sheet_name: str) -> List[Dict]:
        """
        从Excel的压缩文件中提取图片（备用方法）

        Args:
            sheet_name: 工作表名称

        Returns:
            图片信息列表
        """
        images = []

        try:
            # Excel文件本质上是ZIP文件
            with zipfile.ZipFile(self.excel_file_path, 'r') as zip_file:
                # 查找媒体文件
                media_files = [f for f in zip_file.namelist() if f.startswith('xl/media/')]

                for media_file in media_files:
                    try:
                        # 生成随机图片ID
                        image_id = self.generate_random_image_id()

                        # 获取文件扩展名
                        file_ext = Path(media_file).suffix
                        if not file_ext:
                            file_ext = '.png'  # 默认扩展名

                        # 保存图片文件
                        image_filename = f"{image_id}{file_ext}"
                        image_path = (self.images_dir / image_filename).absolute()

                        # 从ZIP中提取图片
                        with zip_file.open(media_file) as source, open(image_path, 'wb') as target:
                            shutil.copyfileobj(source, target)

                        # 记录图片信息
                        image_info = {
                            "image_id": image_id,
                            "filename": image_filename,
                            "sheet_name": sheet_name,
                            "source_file": media_file,
                            "extracted_at": datetime.now().isoformat()
                        }

                        images.append(image_info)

                    except Exception as e:
                        logger.info(f"从ZIP提取图片时出错: {e}")
                        continue

        except Exception as e:
            logger.error(f"从ZIP文件提取图片时出错: {e}")

        return images

    def extract_sheet_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        提取工作表数据

        Args:
            sheet_name: 工作表名称

        Returns:
            工作表数据字典
        """
        try:
            workbook = load_workbook(self.excel_file_path, data_only=True)
            worksheet = workbook[sheet_name]

            sheet_data = {
                "sheet_name": sheet_name,
                "formatted_name": self.get_sheet_naming_convention(sheet_name),
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "cells": {},
                "images": [],
                "extracted_at": datetime.now().isoformat()
            }

            # 提取单元格数据
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_key = f"R{row}C{col}"
                        sheet_data["cells"][cell_key] = {
                            "value": cell.value,
                            "data_type": str(type(cell.value).__name__),
                            "coordinate": cell.coordinate,
                            "row": row,
                            "column": col
                        }

            # 提取图片
            sheet_data["images"] = self.extract_images_from_sheet(sheet_name, sheet_data)

            return sheet_data

        except Exception as e:
            logger.error(f"提取工作表 {sheet_name} 数据时出错: {e}")
            return {
                "sheet_name": sheet_name,
                "formatted_name": self.get_sheet_naming_convention(sheet_name),
                "error": str(e),
                "extracted_at": datetime.now().isoformat()
            }

    def extract_all_data(self) -> Dict[str, Any]:
        """
        提取所有工作表的数据和图片

        Returns:
            包含所有提取数据的字典
        """
        try:
            workbook = load_workbook(self.excel_file_path, data_only=True)
            sheet_names = workbook.sheetnames

            extraction_summary = {
                "excel_file": self.excel_file_path,
                "total_sheets": len(sheet_names),
                "extraction_started_at": datetime.now().isoformat(),
                "sheets": {},
                "summary": {
                    "total_images": 0,
                    "total_cells": 0
                }
            }

            for sheet_name in sheet_names:
                logger.info(f"正在处理工作表: {sheet_name}")
                sheet_data = self.extract_sheet_data(sheet_name)

                # 使用格式化后的名称作为键
                formatted_name = sheet_data["formatted_name"]
                extraction_summary["sheets"][formatted_name] = sheet_data

                # 更新统计信息
                if "images" in sheet_data:
                    extraction_summary["summary"]["total_images"] += len(sheet_data["images"])
                if "cells" in sheet_data:
                    extraction_summary["summary"]["total_cells"] += len(sheet_data["cells"])

            extraction_summary["extraction_completed_at"] = datetime.now().isoformat()

            self.extracted_data = extraction_summary
            return extraction_summary

        except Exception as e:
            logger.error(f"提取Excel数据时出错: {e}")
            return {"error": str(e)}

    def save_to_json(self, filename: str | None = None) -> str:
        """
        将提取的数据保存为JSON文件

        Args:
            filename: 输出文件名（可选）

        Returns:
            保存的文件路径
        """
        if not self.extracted_data:
            logger.info("没有数据可保存，请先运行extract_all_data()")
            return ""

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"excel_extraction_{timestamp}.json"

        json_path = self.data_dir / filename

        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(self.extracted_data, f, ensure_ascii=False, indent=2)

            logger.info(f"数据已保存到: {json_path}")
            return str(json_path)

        except Exception as e:
            logger.error(f"保存JSON文件时出错: {e}")
            return ""

    def generate_report(self) -> str:
        """
        生成提取报告

        Returns:
            报告内容
        """
        if not self.extracted_data:
            return "没有数据可报告"

        report = f"""
Excel数据提取报告
================

文件路径: {self.extracted_data.get('excel_file', 'N/A')}
提取时间: {self.extracted_data.get('extraction_started_at', 'N/A')}
完成时间: {self.extracted_data.get('extraction_completed_at', 'N/A')}

统计信息:
- 总工作表数: {self.extracted_data.get('total_sheets', 0)}
- 总图片数: {self.extracted_data.get('summary', {}).get('total_images', 0)}
- 总单元格数: {self.extracted_data.get('summary', {}).get('total_cells', 0)}

工作表列表:
"""

        for formatted_name, sheet_data in self.extracted_data.get('sheets', {}).items():
            report += f"\n- {formatted_name} (原名称: {sheet_data.get('sheet_name', 'N/A')})"
            if 'images' in sheet_data:
                report += f" - {len(sheet_data['images'])} 张图片"
            if 'cells' in sheet_data:
                report += f" - {len(sheet_data['cells'])} 个单元格"

        return report

    def extract_text_data_to_json(self, sheet_name: str | None = None) -> Dict[str, Any]:
        """
        提取Excel中的文字信息到JSON格式，JSON的key是第一行的列名

        Args:
            sheet_name: 工作表名称，如果为None则处理所有工作表

        Returns:
            包含文字数据的字典
        """
        try:
            workbook = load_workbook(self.excel_file_path, data_only=True)

            if sheet_name:
                # 处理指定工作表
                if sheet_name not in workbook.sheetnames:
                    return {"error": f"工作表 '{sheet_name}' 不存在"}
                sheet_names = [sheet_name]
            else:
                # 处理所有工作表
                sheet_names = workbook.sheetnames

            text_data = {
                "excel_file": self.excel_file_path,
                "extraction_time": datetime.now().isoformat(),
                "sheets": {}
            }

            for current_sheet_name in sheet_names:
                logger.info(f"正在提取工作表 '{current_sheet_name}' 的文字数据...")
                worksheet = workbook[current_sheet_name]

                # 获取第一行作为列名
                headers = []
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    if cell_value is not None:
                        # 清理列名，移除特殊字符，替换空格为下划线
                        clean_header = str(cell_value).strip()
                        clean_header = "".join(c for c in clean_header if c.isalnum() or c in (' ', '-', '_'))
                        clean_header = clean_header.replace(' ', '_').replace('-', '_')
                        if not clean_header:
                            clean_header = f"column_{col}"
                        headers.append(clean_header)
                    else:
                        headers.append(f"column_{col}")

                # 提取数据行
                rows_data = []
                for row in range(2, worksheet.max_row + 1):  # 从第二行开始（跳过标题行）
                    row_data = {}
                    has_data = False

                    for col in range(1, worksheet.max_column + 1):
                        if col <= len(headers):
                            cell_value = worksheet.cell(row=row, column=col).value
                            if cell_value is not None:
                                row_data[headers[col - 1]] = cell_value
                                has_data = True
                            else:
                                row_data[headers[col - 1]] = None

                    # 只添加有数据的行
                    if has_data:
                        rows_data.append(row_data)

                # 保存工作表数据
                text_data["sheets"][current_sheet_name] = {
                    "headers": headers,
                    "total_rows": len(rows_data),
                    "data": rows_data
                }

                logger.info(f"工作表 '{current_sheet_name}' 提取完成: {len(rows_data)} 行数据")

            return text_data

        except Exception as e:
            logger.error(f"提取文字数据时出错: {e}")
            return {"error": str(e)}

    def save_text_data_to_json(self, text_data: Dict[str, Any], filename: str | None = None) -> str:
        """
        将提取的文字数据保存为JSON文件

        Args:
            text_data: 文字数据字典
            filename: 输出文件名（可选）

        Returns:
            保存的文件路径
        """
        if not text_data or "error" in text_data:
            logger.info("没有文字数据可保存")
            return ""

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"excel_text_data_{timestamp}.json"

        json_path = self.data_dir / filename

        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(text_data, f, ensure_ascii=False, indent=2)

            logger.info(f"文字数据已保存到: {json_path}")
            return str(json_path)

        except Exception as e:
            logger.error(f"保存文字数据JSON文件时出错: {e}")
            return ""

    async def parse(self) -> bool:
        """Main parsing method with optimized flow"""
        split_strategy = int(self.file_params.parser.get("split_strategy", SplitStrategy.ROW.value))
        file_repo = KbotMdKbFilesRepository()

        if split_strategy == SplitStrategy.ROW.value:
            # self.chunk_size = int(self.file_params.parser.get("chunk_size", 500))
            # self.chunk_overlap = int(self.file_params.parser.get("chunk_overlap", 50))
            try:
                # Extract all content by page
                all_text_data = self.extract_text_data_to_json()
                if "error" not in all_text_data:
                    logger.info("✓ 成功提取所有工作表数据")
                    logger.info(f"  工作表数: {len(all_text_data['sheets'])}")

                    # 构造最简JSON结构：顶层以sheet名为key，仅保留 total_rows 与 data
                    minimal_data = {}
                    for s_name, s_info in all_text_data['sheets'].items():
                        minimal_data[s_name] = {
                            "data": s_info.get("data", [])
                        }
                    # 保存到JSON文件（只输出 demo_all_sheets.json）
                    json_file = self.save_text_data_to_json(minimal_data, "text_info.json")
                    self.tables_info=minimal_data
                    logger.info(f"  最简数据已保存到: {json_file}")


                    combined_list = []
                    for s_name, s_info in minimal_data.items():
                        rows = s_info.get("data", [])
                        for row in rows:
                            row_str = json.dumps(row, ensure_ascii=False)
                            combined_list.append(f"{s_name}: {row_str}")
                    combined_json_path = self.save_text_data_to_json({"combined": combined_list},
                                                                          "combined_rows_list.json")
                    if combined_json_path:
                        logger.info(f"  合并字符串列表已保存到: {combined_json_path}")
                    # 如果某sheet有图片，抽取并生成 image_info.json（文件名为随机ID，记录 image_id 与 sheet_name）


                    # 显示每个工作表的摘要
                    for sheet_name, sheet_info in all_text_data['sheets'].items():
                        logger.info(f"\n  工作表: {sheet_name}")
                        logger.info(f"    列数: {len(sheet_info['headers'])}")
                        logger.info(f"    数据行数: {sheet_info['total_rows']}")
                        logger.info(
                            f"    列名: {', '.join(sheet_info['headers'][:5])}{'...' if len(sheet_info['headers']) > 5 else ''}")

                        # 显示前几行数据
                        if sheet_info['data']:
                            logger.info(f"    前2行数据预览:")
                            for i, row in enumerate(sheet_info['data'][:2]):
                                # 只显示前3列
                                preview_data = {k: v for k, v in list(row.items())[:3]}
                                logger.info(f"      行{i + 1}: {preview_data}")
                else:
                    logger.info(f"✗ 提取失败: {all_text_data['error']}")
                    return False

                    # 仅输出 demo_all_sheets.json，不进行单表与图片元数据输出

                # Save parsed metadata
                if not await self._process_embeddings():
                    return False
                await file_repo.update_file_parsed_metadata(self.file_params.file_id, self.parsed_metadata)


                return True

            except Exception as e:
                logger.error(f"Error processing Excel file: {str(e)}")
                logger.exception('Error', e)
                await file_repo.update_file_status(
                    self.file_params.file_id,
                    FileStatus.PARSE_FAILED,
                    str(e)
                )
                return False

        else:
            logger.warning(f"Unrecognized split strategy: {split_strategy}")
            return False


async def process_excel(file_params: FileParams) -> bool:
    """
    Process Excel file by extracting content and generating embeddings

    Args:
        file_params: File parameters including path and processing options

    Returns:
        bool: True if processing succeeded, False otherwise
    """
    if not await check_text_file(file_params):
        return False

    try:
        logger.info(f"Processing Excel file: {file_params.file_path}")
        parser = ExcelParser(file_params)
        r = await parser.parse()
        if r:
            msg = f"Successfully parsed {file_params.file_path} (file id: {file_params.file_id})"
            await KbotMdKbFilesRepository().update_file_status(
                file_params.file_id,
                FileStatus.PARSED,
                str(msg)
            )
            return True
        else:
            msg = f"Failed to parse {file_params.file_path} (file id: {file_params.file_id})"
            await KbotMdKbFilesRepository().update_file_status(
                file_params.file_id,
                FileStatus.PARSE_FAILED,
                str(msg)
            )
            return False
    except Exception as e:
        msg = f"Error processing Excel file: {file_params.file_path}, error: {str(e)}"
        logger.error(msg)
        await KbotMdKbFilesRepository().update_file_status(
            file_params.file_id,
            FileStatus.PARSE_FAILED,
            msg
        )
        return False